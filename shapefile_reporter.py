import streamlit as st
import os
import geopandas as gpd
import fiona
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import zipfile
import tempfile
import shutil

# Configuraci√≥n de la p√°gina
st.set_page_config(page_title="Auditor√≠a de Shapefiles", page_icon="üó∫Ô∏è", layout="wide")

def process_shapefiles(root_folder):
    """
    Busca shapefiles recursivamente, extrae metadatos y retorna:
    1. Un objeto BytesIO con el Excel formateado.
    2. El DataFrame generado para visualizaci√≥n.
    3. Una lista de logs/errores (si los hubo).
    """
    filas = []
    logs = []
    
    # Barra de progreso y texto de estado
    status_text = st.empty()
    progress_bar = st.progress(0)
    
    # Primero contamos archivos para la barra de progreso (estimado)
    total_files = sum([len(files) for r, d, files in os.walk(root_folder)])
    processed_count = 0

    for raiz, dirs, archivos in os.walk(root_folder):
        for archivo in archivos:
            processed_count += 1
            # Actualizar progreso cada cierto tiempo para no alentar
            if processed_count % 10 == 0 or processed_count == total_files:
                 progress_bar.progress(min(processed_count / max(total_files, 1), 1.0))
            
            if not archivo.lower().endswith(".shp"):
                continue

            status_text.text(f"Procesando: {archivo}...")
            
            ruta_shp = os.path.join(raiz, archivo)
            carpeta_base = os.path.dirname(ruta_shp)

            tipo_geom = "Unknown"
            props = {}

            # Intento 1: Leer esquema con Fiona
            try:
                with fiona.open(ruta_shp, "r") as src:
                    schema = src.schema
                    props = schema.get("properties", {})
                    tipo_geom = schema.get("geometry", "Unknown")
            except Exception as e:
                logs.append(f"‚ö†Ô∏è Error fiona en {archivo}: {e}")
                continue

            # Intento 2: Refinar datos con Geopandas (opcional pero √∫til para geometr√≠a real)
            gdf = None
            try:
                gdf = gpd.read_file(ruta_shp)
                if not gdf.empty:
                    # Sobrescribir tipo de geometr√≠a con la dominante real
                    tipo_geom = gdf.geom_type.mode()[0]
            except Exception:
                # Si falla geopandas, nos quedamos con lo de fiona
                pass

            # Procesar campos
            if not props:
                # Caso borde: shapefile sin campos
                filas.append({
                    "Carpeta Base": carpeta_base,
                    "Nombre Shapefile": archivo,
                    "Tipo de Geometr√≠a": tipo_geom,
                    "Nombre Campo": "(Sin atributos)",
                    "Tipo de Dato": "",
                    "Longitud": 0
                })
            else:
                for fname, ftype in props.items():
                    tipo = str(ftype)
                    longitud = 0
                    
                    # Parsear tipos tipo "str:80"
                    if isinstance(ftype, str) and ":" in ftype:
                        partes = ftype.split(":", 1)
                        tipo = partes[0]
                        try:
                            longitud = int(partes[1])
                        except:
                            longitud = 0

                    # Calcular longitud real m√°xima si tenemos datos
                    if gdf is not None and fname in gdf.columns:
                        serie = gdf[fname]
                        # Verificar si es tipo texto o object para calcular el len()
                        if pd.api.types.is_object_dtype(serie) or pd.api.types.is_string_dtype(serie):
                            # dropna para evitar error en len(), astype(str) por seguridad
                            serie_no_na = serie.dropna().astype(str)
                            if not serie_no_na.empty:
                                maxlen = int(serie_no_na.map(len).max())
                                # Si la longitud definida es 0, usamos la calculada
                                if longitud == 0:
                                    longitud = maxlen

                    filas.append({
                        "Carpeta Base": carpeta_base,
                        "Nombre Shapefile": archivo,
                        "Tipo de Geometr√≠a": tipo_geom,
                        "Nombre Campo": fname,
                        "Tipo de Dato": tipo,
                        "Longitud": longitud
                    })

    status_text.text("Generando Excel...")
    progress_bar.progress(1.0)

    if not filas:
        return None, pd.DataFrame(), logs

    # === CREAR DATAFRAME ===
    df = pd.DataFrame(filas)
    
    # Ordenar
    df = df.sort_values(by=["Carpeta Base", "Nombre Shapefile", "Nombre Campo"]).reset_index(drop=True)

    # Crear una copia para exportar donde ocultaremos duplicados
    df_export = df.copy()
    df_export["Carpeta Base"] = df_export["Carpeta Base"].mask(df_export["Carpeta Base"].duplicated(), "")
    df_export["Nombre Shapefile"] = df_export["Nombre Shapefile"].mask(df_export["Nombre Shapefile"].duplicated(), "")

    # === EXPORTAR A MEMORIA (BytesIO) ===
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Campos")
        hoja = writer.sheets["Campos"]

        # Ajustar ancho de columnas
        for i, col in enumerate(df_export.columns, 1):
            # Calcular ancho basado en el contenido visible (aproximado)
            # data_len = df_export[col].astype(str).map(len).max()
            # header_len = len(col)
            # max_len = max(data_len, header_len) + 3
            # Nota: openpyxl a veces tiene problemas estimando anchos exactos, 
            # ponemos un ancho razonable por defecto o calculamos simple.
            hoja.column_dimensions[get_column_letter(i)].width = 25

        # === FORMATO VISUAL ===
        bold_font = Font(bold=True, color="000000")
        fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

        # Colorear la primera fila de cada grupo de carpeta
        # La l√≥gica original del script usa la columna 1 ("Carpeta Base") para detectar cambios
        prev_folder = None
        # iteramos desde la fila 2 (datos) hasta el final
        # openpyxl es 1-indexed. row 1 es header. datos empiezan row 2.
        for row in range(2, len(df_export) + 2):
            folder_cell_value = hoja.cell(row=row, column=1).value
            
            # Si hay valor en la columna carpeta (significa que cambi√≥ o es el primero)
            if folder_cell_value and folder_cell_value != prev_folder:
                for col in range(1, len(df_export.columns) + 1):
                    cell = hoja.cell(row=row, column=col)
                    cell.font = bold_font
                    cell.fill = fill
                prev_folder = folder_cell_value

    output.seek(0)
    status_text.empty()
    return output, df, logs

# === INTERFAZ DE USUARIO ===

st.title("üó∫Ô∏è Auditor√≠a de Estructura de Shapefiles")
st.markdown("""
Esta aplicaci√≥n analiza shapefiles contenidos en un archivo ZIP. 
Sube tu archivo ZIP para extraer la estructura (campos, tipos, geometr√≠a) y generar un reporte en Excel.
""")

temp_dir = None
input_folder = None

uploaded_file = st.file_uploader("Sube un archivo .zip que contenga tus shapefiles (incluyendo .shp, .shx, .dbf, etc.)", type="zip")

if uploaded_file is not None:
    process_btn = st.button("üöÄ Procesar ZIP", type="primary")
    
    if process_btn:
        # Crear directorio temporal
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, "upload.zip")
        
        try:
            # Guardar y descomprimir
            with open(zip_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            input_folder = temp_dir
            st.info(f"üìÇ Archivos descomprimidos temporalmente para an√°lisis.")
            
            with st.spinner("Escaneando shapefiles... esto puede tomar un momento."):
                excel_data, df_result, logs = process_shapefiles(input_folder)
            
            if excel_data is None or df_result.empty:
                st.warning("‚ö†Ô∏è No se encontraron archivos .shp en el ZIP subido.")
                if logs:
                    with st.expander("Ver Errores encontrados"):
                        for log in logs:
                            st.write(log)
            else:
                st.success(f"‚úÖ Proceso completado. Se analizaron {len(df_result)} campos.")
                
                # M√©tricas
                cols = st.columns(3)
                cols[0].metric("Total Registros (Campos)", len(df_result))
                cols[1].metric("Shapefiles √önicos", df_result["Nombre Shapefile"].nunique())
                cols[2].metric("Directorios Escaneados", df_result["Carpeta Base"].nunique())

                # Bot√≥n de descarga
                st.download_button(
                    label="üì• Descargar Reporte Excel",
                    data=excel_data,
                    file_name="Reporte_Shapefiles.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                
                # Mostrar datos
                st.subheader("Vista Previa de Datos")
                st.dataframe(df_result, use_container_width=True)
                
                if logs:
                    with st.expander(f"‚ö†Ô∏è Ver Advertencias ({len(logs)})"):
                        for log in logs:
                            st.code(log, language="text")

        except Exception as e:
            st.error(f"‚ùå Ocurri√≥ un error al procesar el archivo ZIP: {e}")
        
        finally:
            # Limpieza de temporales
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
